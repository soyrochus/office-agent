from __future__ import annotations

from copy import copy
from dataclasses import dataclass
from pathlib import Path
import re

from offagent.domain.locators import parse_locator, to_v2_locator
from offagent.domain.models import (
    BlockStyle,
    DocumentRef,
    InlineFragment,
    InlineStyle,
    IndexedItem,
    SectionPayload,
    SheetCell,
    SheetSnapshot,
    StructureSection,
    TextContainerSnapshot,
    VisibleTextRange,
    WorkbookStructure,
    WorksheetSummary,
    XlsxSectionCell,
    XlsxRowEmbedding,
    XlsxRowEmbeddingCell,
)
from offagent.domain.text_fragments import (
    apply_style_to_range,
    fragment_text,
    normalize_fragments,
)
from offagent.errors import InvalidArgumentsError, TargetNotEditableError
from offagent.errors import TargetNotFoundError

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.cell.rich_text import CellRichText, TextBlock
    from openpyxl.cell.text import InlineFont
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils.cell import (
        coordinate_to_tuple,
        get_column_letter,
        range_boundaries,
    )
except ModuleNotFoundError:  # pragma: no cover - exercised through dependency checks
    Workbook = None
    CellRichText = None
    Alignment = None
    Font = None
    InlineFont = None
    PatternFill = None
    TextBlock = None
    load_workbook = None
    coordinate_to_tuple = None
    get_column_letter = None
    range_boundaries = None


@dataclass(frozen=True)
class ResolvedCell:
    sheet_name: str
    coordinate: str
    raw_value: object
    formula: str | None
    display_text: str


class TargetNotAppendableError(TargetNotEditableError):
    """Raised when a requested XLSX target cannot accept append text."""


NUMERIC_TEXT_PATTERN = re.compile(
    r"""
    ^\s*
    [\(\+\-]?
    [$€£]?
    (?:
        \d{1,3}(?:,\d{3})+ |
        \d+
    )
    (?:\.\d+)?
    %?
    \)?
    \s*$
    """,
    re.VERBOSE,
)


def extract_document(document_path: Path) -> list[IndexedItem]:
    workbook = _open_workbook(document_path)
    items: list[IndexedItem] = []

    for worksheet in workbook.worksheets:
        indexed_cells: list[tuple[object, str | None, str]] = []
        row_contexts: dict[int, list[tuple[str, str]]] = {}
        column_contexts: dict[int, list[tuple[str, str]]] = {}

        for row in worksheet.iter_rows():
            for cell in row:
                if not _is_indexable_cell(cell):
                    continue

                formula = _formula_text(cell)
                display_text = _display_text(cell)
                indexed_cells.append((cell, formula, display_text))
                row_contexts.setdefault(cell.row, []).append(
                    (cell.coordinate, display_text)
                )
                column_contexts.setdefault(cell.column, []).append(
                    (cell.coordinate, display_text)
                )

        for cell, formula, display_text in indexed_cells:
            item_id = make_item_id(worksheet.title, cell.coordinate)
            items.append(
                IndexedItem(
                    item_id=item_id,
                    item_type="cell",
                    locator=item_id,
                    preview=display_text[:120],
                    content_text=display_text,
                    metadata={
                        "sheet_name": worksheet.title,
                        "coordinate": cell.coordinate,
                        "raw_value": _metadata_raw_value(cell.value),
                        "formula": formula,
                        "display_text": display_text,
                        "data_type": cell.data_type,
                        "row_context": _context_text(
                            row_contexts[cell.row], exclude=cell.coordinate
                        ),
                        "column_context": _context_text(
                            column_contexts[cell.column],
                            exclude=cell.coordinate,
                        ),
                    },
                )
            )

    return items


def build_embedding_text(item: IndexedItem, document_path: Path) -> str:
    metadata = item.metadata
    return "\n".join(
        [
            f"Workbook: {document_path.name}",
            f"Sheet: {metadata.get('sheet_name', '')}",
            f"Cell: {metadata.get('coordinate', '')}",
            f"Row Context: {metadata.get('row_context', '')}",
            f"Column Context: {metadata.get('column_context', '')}",
            f"Value: {metadata.get('display_text', item.content_text)}",
        ]
    )


def build_row_embeddings(
    items: list[IndexedItem], document_path: Path
) -> list[XlsxRowEmbedding]:
    grouped: dict[tuple[str, int], list[IndexedItem]] = {}

    for item in items:
        if not _is_text_like_item(item):
            continue
        metadata = item.metadata
        coordinate = str(metadata.get("coordinate", ""))
        grouped.setdefault(
            (str(metadata.get("sheet_name", "")), _row_number(coordinate)),
            [],
        ).append(item)

    row_embeddings: list[XlsxRowEmbedding] = []
    for (sheet_name, row_number), row_items in sorted(grouped.items()):
        ordered_items = sorted(
            row_items,
            key=lambda item: _coordinate_sort_key(
                str(item.metadata.get("coordinate", ""))
            ),
        )
        contributing_cells = tuple(
            XlsxRowEmbeddingCell(
                item_id=item.item_id,
                coordinate=str(item.metadata.get("coordinate", "")),
                display_text=str(item.metadata.get("display_text", item.content_text)),
                preview=item.preview,
            )
            for item in ordered_items
        )
        representative = max(
            ordered_items,
            key=lambda item: _representative_score(
                str(item.metadata.get("coordinate", "")), item
            ),
        )
        row_embeddings.append(
            XlsxRowEmbedding(
                sheet_name=sheet_name,
                row_number=row_number,
                text=_build_row_embedding_text(
                    document_path.name,
                    sheet_name,
                    row_number,
                    contributing_cells,
                ),
                preview=representative.preview,
                representative_item_id=representative.item_id,
                contributing_cells=contributing_cells,
            )
        )

    return row_embeddings


def read_cell(document_path: Path, item_id: str) -> str:
    resolved = resolve_cell(document_path, item_id)
    return resolved.display_text


def create_xlsx(output_path: Path, initial_sheet_name: str | None = None) -> Path:
    if Workbook is None:
        raise RuntimeError("openpyxl is required for XLSX operations.")
    workbook = Workbook()
    workbook.active.title = initial_sheet_name or "Sheet1"
    workbook.save(output_path)
    return output_path


def add_sheet(
    document_path: Path,
    name: str,
    output_path: Path | None = None,
) -> tuple[Path, str]:
    workbook = _open_workbook(document_path)
    if not name.strip():
        raise InvalidArgumentsError("Worksheet name cannot be empty.")
    if name in workbook.sheetnames:
        raise InvalidArgumentsError(f"Worksheet {name!r} already exists.")
    workbook.create_sheet(title=name)
    target_path = _target_path(document_path, output_path)
    workbook.save(target_path)
    return target_path, f"xlsx:sheet:{name}"


def write_cell(
    document_path: Path, item_id: str, value: object, output_path: Path | None = None
) -> Path:
    workbook = _open_workbook(document_path)
    cell = _resolve_cell(workbook, item_id)
    cell.value = _coerce_write_value(value)
    target_path = _target_path(document_path, output_path)
    workbook.save(target_path)
    return target_path


def add_row(
    document_path: Path,
    sheet_locator: str,
    values: list[object],
    output_path: Path | None = None,
) -> tuple[Path, str]:
    workbook = _open_workbook(document_path)
    sheet_name = _sheet_name_from_locator(sheet_locator)
    worksheet = _resolve_worksheet(workbook, sheet_name)
    target_row = _last_used_row(worksheet) + 1
    for column_index, value in enumerate(values, start=1):
        worksheet.cell(row=target_row, column=column_index).value = _coerce_write_value(
            value
        )
    target_path = _target_path(document_path, output_path)
    workbook.save(target_path)
    return target_path, f"xlsx:sheet:{sheet_name}:row:{target_row}"


def append_cell(
    document_path: Path, item_id: str, text: str, output_path: Path | None = None
) -> Path:
    workbook = _open_workbook(document_path)
    cell = _resolve_cell(workbook, item_id)
    formula = _formula_text(cell)
    if formula is not None or (
        cell.value is not None and not isinstance(cell.value, str)
    ):
        raise TargetNotAppendableError("target not appendable; use write-cell")
    if cell.value is None:
        cell.value = text
    else:
        cell.value = f"{cell.value}{text}"
    target_path = _target_path(document_path, output_path)
    workbook.save(target_path)
    return target_path


def resolve_cell(document_path: Path, item_id: str) -> ResolvedCell:
    workbook = _open_workbook(document_path)
    cell = _resolve_cell(workbook, item_id)
    return ResolvedCell(
        sheet_name=cell.parent.title,
        coordinate=cell.coordinate,
        raw_value=cell.value,
        formula=_formula_text(cell),
        display_text=_display_text(cell),
    )


def resolve_structure(document_path: Path) -> tuple[StructureSection, ...]:
    workbook = _open_workbook(document_path)
    sections: list[StructureSection] = []

    for worksheet in workbook.worksheets:
        anchor = _first_indexable_cell(worksheet)
        locator = make_item_id(
            worksheet.title, anchor.coordinate if anchor is not None else "A1"
        )
        preview = "" if anchor is None else _display_text(anchor)[:120]
        sections.append(
            StructureSection(
                locator=locator,
                section_type="worksheet",
                preview=preview,
                metadata={
                    "sheet_name": worksheet.title,
                    "used_range": _format_range(_used_bounds(worksheet)),
                    "max_row": worksheet.max_row,
                    "max_column": worksheet.max_column,
                    "cell_count": _cell_count(worksheet),
                },
            )
        )

    return tuple(sections)


def get_section(
    document_path: Path, locator: str, *, cell_range: str | None = None
) -> SectionPayload:
    sheet_name, _ = parse_item_id(locator)
    snapshot = get_sheet_snapshot(document_path, sheet_name, cell_range=cell_range)
    return SectionPayload(
        document=snapshot.document,
        locator=locator,
        section_type="worksheet",
        preview=next(
            (cell.display_value for cell in snapshot.cells if cell.display_value), ""
        ),
        metadata={**snapshot.metadata, "sheet_name": sheet_name},
        sheet_name=sheet_name,
        cells=tuple(
            XlsxSectionCell(
                locator=make_item_id(sheet_name, cell.coordinate),
                coordinate=cell.coordinate,
                row=cell.row,
                column=cell.column,
                display_value=cell.display_value,
                formula=cell.metadata.get("formula"),
                metadata=cell.metadata,
            )
            for cell in snapshot.cells
        ),
    )


def read_node(document_path: Path, locator: str) -> tuple[str, str, dict[str, object]]:
    resolved = resolve_cell(document_path, locator)
    return (
        "cell",
        resolved.display_text,
        {
            "sheet_name": resolved.sheet_name,
            "coordinate": resolved.coordinate,
            "formula": resolved.formula,
            "raw_value": resolved.raw_value,
        },
    )


def write_node(
    document_path: Path, locator: str, value: str, output_path: Path | None = None
) -> Path:
    return write_cell(document_path, locator, value, output_path)


def get_workbook_structure(document_path: Path) -> WorkbookStructure:
    workbook = _open_workbook(document_path)
    sheets: list[WorksheetSummary] = []

    for position, worksheet in enumerate(workbook.worksheets):
        used_bounds = _used_bounds(worksheet)
        preview = ""
        if used_bounds is not None:
            min_row, min_col, max_row, max_col = used_bounds
            for row in worksheet.iter_rows(
                min_row=min_row,
                max_row=max_row,
                min_col=min_col,
                max_col=max_col,
            ):
                for cell in row:
                    text = _display_text(cell).strip()
                    if text:
                        preview = text[:120]
                        break
                if preview:
                    break
        sheets.append(
            WorksheetSummary(
                position=position,
                sheet_name=worksheet.title,
                preview=preview,
                metadata={
                    "used_range": _format_range(used_bounds),
                    "max_row": worksheet.max_row,
                    "max_column": worksheet.max_column,
                },
            )
        )

    return WorkbookStructure(
        document=_document_ref(document_path), sheets=tuple(sheets)
    )


def get_sheet_snapshot(
    document_path: Path,
    sheet_name: str,
    *,
    cell_range: str | None = None,
    start_cell: str | None = None,
    row_count: int | None = None,
    column_count: int | None = None,
) -> SheetSnapshot:
    workbook = _open_workbook(document_path)
    worksheet = _resolve_worksheet(workbook, sheet_name)
    bounds = _snapshot_bounds(
        worksheet,
        cell_range=cell_range,
        start_cell=start_cell,
        row_count=row_count,
        column_count=column_count,
    )

    cells: list[SheetCell] = []
    if bounds is not None:
        min_row, min_col, max_row, max_col = bounds
        for row in worksheet.iter_rows(
            min_row=min_row,
            max_row=max_row,
            min_col=min_col,
            max_col=max_col,
        ):
            for cell in row:
                cells.append(
                    SheetCell(
                        coordinate=cell.coordinate,
                        row=cell.row,
                        column=cell.column,
                        display_value=_display_text(cell),
                        metadata={
                            "raw_value": _metadata_raw_value(cell.value),
                            "formula": _formula_text(cell),
                            "data_type": cell.data_type,
                        },
                    )
                )

    return SheetSnapshot(
        document=_document_ref(document_path),
        sheet_name=worksheet.title,
        cells=tuple(cells),
        metadata={
            "range": _format_range(bounds),
            "row_count": 0 if bounds is None else bounds[2] - bounds[0] + 1,
            "column_count": 0 if bounds is None else bounds[3] - bounds[1] + 1,
        },
    )


def append_row(
    document_path: Path,
    sheet_name: str,
    *,
    values: list[object] | None = None,
    record: dict[str, object] | None = None,
    output_path: Path | None = None,
) -> tuple[Path, int, tuple[str, ...]]:
    if (values is None) == (record is None):
        raise InvalidArgumentsError(
            "append_row requires exactly one of values or record."
        )

    workbook = _open_workbook(document_path)
    worksheet = _resolve_worksheet(workbook, sheet_name)
    target_row = _last_used_row(worksheet) + 1
    written_coordinates: list[str] = []

    if values is not None:
        for column_index, value in enumerate(values, start=1):
            coordinate = f"{get_column_letter(column_index)}{target_row}"
            worksheet[coordinate] = _coerce_write_value(value)
            written_coordinates.append(coordinate)
    else:
        header_map = _header_map(worksheet)
        if not header_map:
            raise InvalidArgumentsError(
                "append_row record writes require an existing header row in the worksheet."
            )
        for key, value in record.items():
            if key not in header_map:
                raise InvalidArgumentsError(
                    f"Unknown worksheet header for append_row: {key}"
                )
            coordinate = f"{get_column_letter(header_map[key])}{target_row}"
            worksheet[coordinate] = _coerce_write_value(value)
            written_coordinates.append(coordinate)

    target_path = _target_path(document_path, output_path)
    workbook.save(target_path)
    return target_path, target_row, tuple(written_coordinates)


def write_table(
    document_path: Path,
    sheet_name: str,
    *,
    rows: list[list[object]] | None = None,
    records: list[dict[str, object]] | None = None,
    column_mapping: dict[str, str] | None = None,
    output_path: Path | None = None,
) -> tuple[Path, int, int]:
    if (rows is None) == (records is None):
        raise InvalidArgumentsError(
            "write_table requires exactly one of rows or records."
        )

    workbook = _open_workbook(document_path)
    worksheet = _resolve_worksheet(workbook, sheet_name)
    start_row = _last_used_row(worksheet) + 1

    if rows is not None:
        for row_offset, row_values in enumerate(rows):
            for column_index, value in enumerate(row_values, start=1):
                worksheet.cell(
                    row=start_row + row_offset,
                    column=column_index,
                ).value = _coerce_write_value(value)
        end_row = start_row + len(rows) - 1
    else:
        resolved_mapping = _resolve_record_mapping(worksheet, column_mapping)
        for row_offset, record in enumerate(records):
            for key, value in record.items():
                if key not in resolved_mapping:
                    raise InvalidArgumentsError(
                        f"Unknown worksheet mapping for write_table field: {key}"
                    )
                worksheet.cell(
                    row=start_row + row_offset,
                    column=resolved_mapping[key],
                ).value = _coerce_write_value(value)
        end_row = start_row + len(records) - 1

    target_path = _target_path(document_path, output_path)
    workbook.save(target_path)
    return target_path, start_row, end_row


def parse_item_id(item_id: str) -> tuple[str, str]:
    if not item_id.startswith("sheet:"):
        raise InvalidArgumentsError(f"Unsupported XLSX item id: {item_id}")

    payload = item_id.removeprefix("sheet:")
    if "!" not in payload:
        raise InvalidArgumentsError(f"Invalid XLSX item id: {item_id}")

    sheet_name, coordinate = payload.rsplit("!", maxsplit=1)
    if not sheet_name:
        raise InvalidArgumentsError(f"Invalid XLSX sheet name in item id: {item_id}")
    normalized_coordinate = _normalize_coordinate(coordinate)
    return sheet_name, normalized_coordinate


def make_item_id(sheet_name: str, coordinate: str) -> str:
    return f"sheet:{sheet_name}!{_normalize_coordinate(coordinate)}"


def style_cell_inline(
    document_path: Path,
    locator: str,
    style: InlineStyle,
    clear_fields: list[str] | tuple[str, ...],
    output_path: Path | None = None,
) -> tuple[Path, str, dict[str, object]]:
    workbook = _open_workbook(document_path)
    canonical = to_v2_locator(locator, file_type="xlsx")
    cell = _resolve_cell(workbook, _legacy_item_id_from_v2(canonical))
    clear_set = _normalize_clear_fields(clear_fields, _INLINE_STYLE_FIELDS)
    skipped_fields = _apply_xlsx_inline_style(cell, style, clear_set)
    target_path = _target_path(document_path, output_path)
    workbook.save(target_path)
    return (
        target_path,
        canonical,
        {"cleared_fields": clear_set, "skipped_fields": skipped_fields},
    )


def read_cell_fragments(document_path: Path, locator: str) -> TextContainerSnapshot:
    workbook = _open_workbook(document_path)
    canonical = to_v2_locator(locator, file_type="xlsx")
    cell = _resolve_cell(workbook, _legacy_item_id_from_v2(canonical))
    _ensure_partial_formatting_cell_supported(cell, canonical)
    fragments = _read_xlsx_fragments(cell)
    return TextContainerSnapshot(
        locator=canonical,
        object_type="cell",
        text=fragment_text(fragments),
        fragments=fragments,
        metadata={"sheet_name": cell.parent.title, "coordinate": cell.coordinate},
    )


def write_cell_fragments(
    document_path: Path,
    locator: str,
    fragments: list[InlineFragment] | tuple[InlineFragment, ...],
    output_path: Path | None = None,
) -> tuple[Path, str, TextContainerSnapshot]:
    workbook = _open_workbook(document_path)
    canonical = to_v2_locator(locator, file_type="xlsx")
    cell = _resolve_cell(workbook, _legacy_item_id_from_v2(canonical))
    _ensure_partial_formatting_cell_supported(cell, canonical)
    normalized = normalize_fragments(fragments)
    _write_xlsx_fragments(cell, normalized)
    target_path = _target_path(document_path, output_path)
    workbook.save(target_path)
    snapshot = TextContainerSnapshot(
        locator=canonical,
        object_type="cell",
        text=fragment_text(normalized),
        fragments=normalized,
        metadata={"sheet_name": cell.parent.title, "coordinate": cell.coordinate},
    )
    return target_path, canonical, snapshot


def style_cell_range(
    document_path: Path,
    locator: str,
    text_range: VisibleTextRange,
    style: InlineStyle,
    clear_fields: list[str] | tuple[str, ...],
    output_path: Path | None = None,
) -> tuple[Path, str, dict[str, object]]:
    snapshot = read_cell_fragments(document_path, locator)
    clear_set = _normalize_clear_fields(clear_fields, _INLINE_STYLE_FIELDS)
    styled = apply_style_to_range(
        snapshot.fragments, text_range, style=style, clear_fields=clear_set
    )
    target_path, canonical, rewritten = write_cell_fragments(
        document_path,
        locator,
        styled,
        output_path=output_path,
    )
    return (
        target_path,
        canonical,
        {
            "cleared_fields": clear_set,
            "range": {"start": text_range.start, "end": text_range.end},
            "text": rewritten.text,
        },
    )


def style_cell_block(
    document_path: Path,
    locator: str,
    style: BlockStyle,
    clear_fields: list[str] | tuple[str, ...],
    output_path: Path | None = None,
) -> tuple[Path, str, dict[str, object]]:
    workbook = _open_workbook(document_path)
    canonical = to_v2_locator(locator, file_type="xlsx")
    cell = _resolve_cell(workbook, _legacy_item_id_from_v2(canonical))
    clear_set = _normalize_clear_fields(clear_fields, _BLOCK_STYLE_FIELDS)
    skipped_fields = _apply_xlsx_block_style(cell, style, clear_set)
    target_path = _target_path(document_path, output_path)
    workbook.save(target_path)
    return (
        target_path,
        canonical,
        {"cleared_fields": clear_set, "skipped_fields": skipped_fields},
    )


def _open_workbook(document_path: Path):
    if load_workbook is None:
        raise RuntimeError("openpyxl is required for XLSX operations.")
    return load_workbook(str(document_path), rich_text=True)


def _document_ref(document_path: Path) -> DocumentRef:
    resolved_path = document_path.resolve()
    stat = resolved_path.stat()
    return DocumentRef(
        document_id=resolved_path.as_posix(),
        path=resolved_path,
        file_type="xlsx",
        display_name=resolved_path.name,
        modified_time=stat.st_mtime,
    )


def _resolve_cell(workbook, item_id: str):
    sheet_name, coordinate = parse_item_id(item_id)
    worksheet = _resolve_worksheet(workbook, sheet_name)
    return worksheet[coordinate]


def _resolve_worksheet(workbook, sheet_name: str):
    try:
        return workbook[sheet_name]
    except KeyError as exc:
        raise TargetNotFoundError(
            f"Worksheet {sheet_name!r} does not exist in the workbook."
        ) from exc


def _is_indexable_cell(cell) -> bool:
    return _formula_text(cell) is not None or cell.value is not None


def _formula_text(cell) -> str | None:
    if getattr(cell, "data_type", None) == "f" and cell.value is not None:
        return str(cell.value)
    return None


def _display_text(cell) -> str:
    formula = _formula_text(cell)
    if formula is not None:
        return formula
    return "" if cell.value is None else str(cell.value)


def _metadata_raw_value(value: object) -> object:
    if CellRichText is not None and isinstance(value, CellRichText):
        return str(value)
    return value


def _coerce_value(value: str) -> object:
    for converter in (int, float):
        try:
            return converter(value)
        except ValueError:
            continue
    return value


def _coerce_write_value(value: object) -> object:
    if isinstance(value, str):
        return _coerce_value(value)
    return value


def _normalize_coordinate(coordinate: str) -> str:
    normalized = coordinate.strip().upper()
    if not normalized:
        raise InvalidArgumentsError("Cell coordinate cannot be empty.")
    if coordinate_to_tuple is None:
        raise RuntimeError("openpyxl is required for XLSX operations.")
    try:
        coordinate_to_tuple(normalized)
    except ValueError as exc:
        raise InvalidArgumentsError(
            f"Invalid XLSX cell coordinate: {coordinate}"
        ) from exc
    return normalized


def _context_text(entries: list[tuple[str, str]], *, exclude: str) -> str:
    return " | ".join(
        display_text
        for coordinate, display_text in entries
        if coordinate != exclude and display_text
    )


def _target_path(document_path: Path, output_path: Path | None) -> Path:
    return document_path if output_path is None else output_path


def _sheet_name_from_locator(locator: str) -> str:
    canonical = to_v2_locator(locator, file_type="xlsx")
    components = parse_locator(canonical).components
    if len(components) == 3 and components[:2] == ("xlsx", "sheet"):
        return components[2]
    raise InvalidArgumentsError(f"Unsupported worksheet locator: {locator}")


def _legacy_item_id_from_v2(locator: str) -> str:
    components = parse_locator(locator).components
    if len(components) == 4 and components[:2] == ("xlsx", "sheet"):
        return make_item_id(components[2], components[3])
    raise InvalidArgumentsError(f"XLSX cell locator required: {locator}")


def _ensure_partial_formatting_cell_supported(cell, locator: str) -> None:
    if _formula_text(cell) is not None:
        raise TargetNotEditableError(
            f"{locator} does not support partial formatting for formula cells."
        )
    if isinstance(cell.value, bool):
        raise TargetNotEditableError(
            f"{locator} does not support partial formatting for boolean cells."
        )
    if cell.coordinate in {
        merged.split(":")[0] for merged in map(str, cell.parent.merged_cells.ranges)
    }:
        return
    for merged in cell.parent.merged_cells.ranges:
        if cell.coordinate in merged:
            raise TargetNotEditableError(
                f"{locator} does not support partial formatting for merged cells."
            )
    if cell.value is None:
        return
    if CellRichText is not None and isinstance(cell.value, CellRichText):
        return
    if not isinstance(cell.value, str):
        raise TargetNotEditableError(
            f"{locator} does not support partial formatting for non-string cells."
        )


def _read_xlsx_fragments(cell) -> tuple[InlineFragment, ...]:
    value = cell.value
    if value is None:
        return ()
    if CellRichText is not None and isinstance(value, CellRichText):
        fragments: list[InlineFragment] = []
        for part in value:
            if isinstance(part, str):
                fragments.append(InlineFragment(text=part, style=InlineStyle()))
                continue
            fragments.append(
                InlineFragment(
                    text=part.text,
                    style=_inline_style_from_xlsx_font(part.font),
                )
            )
        return normalize_fragments(fragments)
    return (InlineFragment(text=str(value), style=InlineStyle()),)


def _write_xlsx_fragments(
    cell,
    fragments: list[InlineFragment] | tuple[InlineFragment, ...],
) -> None:
    normalized = normalize_fragments(fragments)
    if not normalized:
        cell.value = ""
        return
    if CellRichText is None or TextBlock is None or InlineFont is None:
        raise RuntimeError(
            "openpyxl rich-text support is required for XLSX partial formatting."
        )
    rich_parts: list[object] = []
    for fragment in normalized:
        if all(value is None for value in fragment.style.__dict__.values()):
            rich_parts.append(fragment.text)
            continue
        rich_parts.append(TextBlock(_xlsx_inline_font(fragment.style), fragment.text))
    cell.value = CellRichText(*rich_parts)


def _inline_style_from_xlsx_font(font) -> InlineStyle:
    if font is None:
        return InlineStyle()
    color = None
    if getattr(font, "color", None) not in {None, ""}:
        if isinstance(font.color, str):
            color = font.color[-6:]
        elif getattr(font.color, "rgb", None):
            color = str(font.color.rgb)[-6:]
    underline = None
    if getattr(font, "u", None) is not None:
        underline = str(font.u).lower() not in {"", "none"}
    return InlineStyle(
        bold=getattr(font, "b", getattr(font, "bold", None)),
        italic=getattr(font, "i", getattr(font, "italic", None)),
        underline=underline
        if underline is not None
        else getattr(font, "underline", None),
        strike=getattr(font, "strike", None),
        font_name=getattr(font, "rFont", getattr(font, "name", None)),
        font_size=getattr(font, "sz", None),
        font_color=color,
    )


def _xlsx_inline_font(style: InlineStyle):
    return InlineFont(
        b=style.bold,
        i=style.italic,
        strike=style.strike,
        rFont=style.font_name,
        sz=style.font_size,
        color=None
        if style.font_color is None
        else _normalize_hex_color(style.font_color),
        u="single" if style.underline else None,
    )


_INLINE_STYLE_FIELDS = frozenset(
    {
        "bold",
        "italic",
        "underline",
        "strike",
        "font_name",
        "font_size",
        "font_color",
        "highlight",
    }
)
_BLOCK_STYLE_FIELDS = frozenset(
    {
        "alignment",
        "indent_level",
        "left_indent",
        "right_indent",
        "spacing_before",
        "spacing_after",
        "line_spacing",
        "wrap_text",
        "vertical_alignment",
        "fill_color",
        "number_format",
    }
)
_XLSX_ALIGNMENT_MAP = {
    "left": "left",
    "center": "center",
    "right": "right",
    "justify": "justify",
}
_XLSX_VERTICAL_ALIGNMENT_MAP = {
    "top": "top",
    "center": "center",
    "bottom": "bottom",
}


def _normalize_clear_fields(
    clear_fields: list[str] | tuple[str, ...],
    allowed: frozenset[str],
) -> tuple[str, ...]:
    normalized: list[str] = []
    seen: set[str] = set()
    for field_name in clear_fields:
        if field_name not in allowed:
            raise InvalidArgumentsError(
                f"Unknown style field in clear_fields: {field_name}"
            )
        if field_name not in seen:
            normalized.append(field_name)
            seen.add(field_name)
    return tuple(normalized)


def _apply_xlsx_inline_style(
    cell, style: InlineStyle, clear_fields: tuple[str, ...]
) -> list[str]:
    font = copy(cell.font)
    clear_set = set(clear_fields)
    skipped_fields: list[str] = []

    if "bold" in clear_set:
        font.bold = None
    elif style.bold is not None:
        font.bold = style.bold

    if "italic" in clear_set:
        font.italic = None
    elif style.italic is not None:
        font.italic = style.italic

    if "underline" in clear_set:
        font.underline = None
    elif style.underline is not None:
        font.underline = "single" if style.underline else None

    if "strike" in clear_set:
        font.strike = None
    elif style.strike is not None:
        font.strike = style.strike

    if "font_name" in clear_set:
        font.name = None
    elif style.font_name is not None:
        font.name = style.font_name

    if "font_size" in clear_set:
        font.sz = None
    elif style.font_size is not None:
        font.sz = style.font_size

    if "font_color" in clear_set:
        font.color = None
    elif style.font_color is not None:
        font.color = _normalize_hex_color(style.font_color)

    if style.highlight is not None or "highlight" in clear_set:
        skipped_fields.append("highlight")

    cell.font = font
    return skipped_fields


def _apply_xlsx_block_style(
    cell, style: BlockStyle, clear_fields: tuple[str, ...]
) -> list[str]:
    alignment = copy(cell.alignment)
    clear_set = set(clear_fields)
    skipped_fields: list[str] = []

    if "alignment" in clear_set:
        alignment.horizontal = None
    elif style.alignment is not None:
        alignment.horizontal = _xlsx_alignment_value(style.alignment)

    if "wrap_text" in clear_set:
        alignment.wrap_text = None
    elif style.wrap_text is not None:
        alignment.wrap_text = style.wrap_text

    if "vertical_alignment" in clear_set:
        alignment.vertical = None
    elif style.vertical_alignment is not None:
        alignment.vertical = _xlsx_vertical_alignment_value(style.vertical_alignment)

    if "indent_level" in clear_set:
        alignment.indent = 0
    elif style.indent_level is not None:
        alignment.indent = style.indent_level

    for field_name in (
        "left_indent",
        "right_indent",
        "spacing_before",
        "spacing_after",
        "line_spacing",
    ):
        if getattr(style, field_name) is not None or field_name in clear_set:
            skipped_fields.append(field_name)

    if "fill_color" in clear_set:
        cell.fill = PatternFill(fill_type=None)
    elif style.fill_color is not None:
        color = _normalize_hex_color(style.fill_color)
        cell.fill = PatternFill(fill_type="solid", start_color=color, end_color=color)

    if "number_format" in clear_set:
        cell.number_format = "General"
    elif style.number_format is not None:
        cell.number_format = style.number_format

    cell.alignment = alignment
    return skipped_fields


def _normalize_hex_color(value: str) -> str:
    normalized = value.strip().lstrip("#").upper()
    if len(normalized) != 6 or any(
        character not in "0123456789ABCDEF" for character in normalized
    ):
        raise InvalidArgumentsError(f"Invalid RGB hex color: {value}")
    return normalized


def _xlsx_alignment_value(raw: str) -> str:
    normalized = raw.strip().lower()
    if normalized not in _XLSX_ALIGNMENT_MAP:
        raise InvalidArgumentsError(f"Unsupported XLSX alignment: {raw}")
    return _XLSX_ALIGNMENT_MAP[normalized]


def _xlsx_vertical_alignment_value(raw: str) -> str:
    normalized = raw.strip().lower()
    if normalized not in _XLSX_VERTICAL_ALIGNMENT_MAP:
        raise InvalidArgumentsError(f"Unsupported XLSX vertical alignment: {raw}")
    return _XLSX_VERTICAL_ALIGNMENT_MAP[normalized]


def _is_text_like_item(item: IndexedItem) -> bool:
    metadata = item.metadata
    display_text = str(metadata.get("display_text", item.content_text)).strip()
    if not display_text:
        return False

    raw_value = metadata.get("raw_value")
    if metadata.get("formula") is not None:
        return False
    if isinstance(raw_value, bool):
        return False
    if isinstance(raw_value, (int, float)):
        return False
    if NUMERIC_TEXT_PATTERN.match(display_text):
        return False
    return any(character.isalpha() for character in display_text)


def _row_number(coordinate: str) -> int:
    row_number, _ = _coordinate_sort_key(coordinate)
    return row_number


def _coordinate_sort_key(coordinate: str) -> tuple[int, int]:
    normalized = _normalize_coordinate(coordinate)
    if coordinate_to_tuple is None:
        raise RuntimeError("openpyxl is required for XLSX operations.")
    return coordinate_to_tuple(normalized)


def _representative_score(coordinate: str, item: IndexedItem) -> tuple[int, int, int]:
    display_text = str(item.metadata.get("display_text", item.content_text))
    _, column_number = _coordinate_sort_key(coordinate)
    alpha_characters = sum(1 for character in display_text if character.isalpha())
    return (alpha_characters, len(display_text), -column_number)


def _build_row_embedding_text(
    workbook_name: str,
    sheet_name: str,
    row_number: int,
    contributing_cells: tuple[XlsxRowEmbeddingCell, ...],
) -> str:
    lines = [
        f"Workbook: {workbook_name}",
        f"Sheet: {sheet_name}",
        f"Row: {row_number}",
        "Cells:",
    ]
    lines.extend(
        f"- {cell.coordinate}: {cell.display_text}" for cell in contributing_cells
    )
    return "\n".join(lines)


def _used_bounds(worksheet) -> tuple[int, int, int, int] | None:
    used_cells = [
        cell
        for row in worksheet.iter_rows()
        for cell in row
        if _is_indexable_cell(cell)
    ]
    if not used_cells:
        return None
    min_row = min(cell.row for cell in used_cells)
    min_col = min(cell.column for cell in used_cells)
    max_row = max(cell.row for cell in used_cells)
    max_col = max(cell.column for cell in used_cells)
    return (min_row, min_col, max_row, max_col)


def _snapshot_bounds(
    worksheet,
    *,
    cell_range: str | None,
    start_cell: str | None,
    row_count: int | None,
    column_count: int | None,
) -> tuple[int, int, int, int] | None:
    if cell_range is not None:
        if start_cell is not None or row_count is not None or column_count is not None:
            raise InvalidArgumentsError(
                "sheet snapshot range and window inputs are mutually exclusive."
            )
        if range_boundaries is None:
            raise RuntimeError("openpyxl is required for XLSX operations.")
        min_col, min_row, max_col, max_row = range_boundaries(cell_range)
        return (min_row, min_col, max_row, max_col)

    if start_cell is not None or row_count is not None or column_count is not None:
        if start_cell is None or row_count is None or column_count is None:
            raise InvalidArgumentsError(
                "sheet snapshot windows require start_cell, row_count, and column_count together."
            )
        start_row, start_column = _coordinate_sort_key(start_cell)
        return (
            start_row,
            start_column,
            start_row + row_count - 1,
            start_column + column_count - 1,
        )

    return _used_bounds(worksheet)


def _format_range(bounds: tuple[int, int, int, int] | None) -> str | None:
    if bounds is None or get_column_letter is None:
        return None
    min_row, min_col, max_row, max_col = bounds
    return (
        f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"
    )


def _last_used_row(worksheet) -> int:
    bounds = _used_bounds(worksheet)
    return 0 if bounds is None else bounds[2]


def _header_map(worksheet) -> dict[str, int]:
    header_map: dict[str, int] = {}
    for cell in worksheet[1]:
        header = _display_text(cell).strip()
        if header:
            header_map[header] = cell.column
    return header_map


def _resolve_record_mapping(
    worksheet, column_mapping: dict[str, str] | None
) -> dict[str, int]:
    if column_mapping is None:
        return _header_map(worksheet)

    header_map = _header_map(worksheet)
    resolved: dict[str, int] = {}
    for field_name, target in column_mapping.items():
        normalized_target = target.strip()
        if normalized_target in header_map:
            resolved[field_name] = header_map[normalized_target]
            continue
        if _is_column_reference(normalized_target):
            if coordinate_to_tuple is None:
                raise RuntimeError("openpyxl is required for XLSX operations.")
            _, column_number = coordinate_to_tuple(f"{normalized_target.upper()}1")
            resolved[field_name] = column_number
            continue
        raise InvalidArgumentsError(
            f"Unknown worksheet header in column_mapping: {target}"
        )
    return resolved


def _is_column_reference(value: str) -> bool:
    return bool(re.fullmatch(r"[A-Za-z]+", value))


def _first_indexable_cell(worksheet):
    for row in worksheet.iter_rows():
        for cell in row:
            if _is_indexable_cell(cell):
                return cell
    return None


def _cell_count(worksheet) -> int:
    return sum(
        1 for row in worksheet.iter_rows() for cell in row if _is_indexable_cell(cell)
    )
