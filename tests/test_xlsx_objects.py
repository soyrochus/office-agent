from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.table import Table, TableStyleInfo

from offagent.domain.models import Capability
from offagent.errors import InvalidArgumentsError
from offagent.objects.xlsx_objects import (
    XlsxObjectResolver,
    insert_columns,
    insert_rows,
    merge_cells,
    set_formula,
    write_range,
)


@pytest.fixture
def xlsx_with_object_features(tmp_path) -> Path:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Data"
    worksheet["A1"] = "Name"
    worksheet["B1"] = "Value"
    worksheet["A2"] = "Alpha"
    worksheet["B2"] = 10
    worksheet["C3"] = "=SUM(B2,5)"
    worksheet["D1"] = "Merged"
    worksheet.merge_cells("D1:E1")

    table = Table(displayName="SalesTable", ref="A1:B2")
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    worksheet.add_table(table)

    workbook.defined_names.add(DefinedName("NamedTotals", attr_text="Data!$A$1:$B$2"))
    workbook.create_sheet("Archive")

    path = tmp_path / "objects.xlsx"
    workbook.save(path)
    return path


def test_get_workbook_worksheet_row_column_and_cell_objects(sample_xlsx) -> None:
    resolver = XlsxObjectResolver()

    workbook = resolver.get_object(sample_xlsx, "xlsx:workbook")
    worksheet = resolver.get_object(sample_xlsx, "xlsx:sheet:Budget2026")
    row = resolver.get_object(sample_xlsx, "xlsx:sheet:Budget2026:row:1")
    column = resolver.get_object(sample_xlsx, "xlsx:sheet:Budget2026:col:2")
    cell = resolver.get_object(sample_xlsx, "sheet:Budget2026!B2")

    assert workbook.object_type == "workbook"
    assert workbook.properties["sheet_count"] == 2
    assert worksheet.object_type == "worksheet"
    assert worksheet.properties["sheet_name"] == "Budget2026"
    assert row.object_type == "row"
    assert column.object_type == "column"
    assert cell.object_type == "cell"
    assert cell.properties["coordinate"] == "B2"


def test_get_range_table_merged_formula_and_named_range_objects(xlsx_with_object_features) -> None:
    resolver = XlsxObjectResolver()

    cell_range = resolver.get_object(xlsx_with_object_features, "xlsx:sheet:Data!A1:B2")
    table = resolver.get_object(xlsx_with_object_features, "xlsx:sheet:Data:table:SalesTable")
    merged = resolver.get_object(xlsx_with_object_features, "xlsx:sheet:Data:merged_range:D1:E1")
    formula = resolver.get_object(xlsx_with_object_features, "xlsx:sheet:Data:formula_cell:C3")
    named_range = resolver.get_object(xlsx_with_object_features, "xlsx:named_range:NamedTotals")

    assert cell_range.object_type == "range"
    assert cell_range.properties["cell_count"] == 4
    assert table.object_type == "table"
    assert table.properties["table_name"] == "SalesTable"
    assert merged.object_type == "merged_range"
    assert merged.properties["range"] == "D1:E1"
    assert formula.object_type == "formula_cell"
    assert formula.properties["formula"] == "=SUM(B2,5)"
    assert named_range.object_type == "named_range"
    assert "Data!" in named_range.properties["reference"]


def test_list_children_filters_worksheet_range_and_table_children(xlsx_with_object_features) -> None:
    resolver = XlsxObjectResolver()

    workbook_children = resolver.list_children(xlsx_with_object_features, "xlsx:workbook", child_type="worksheet")
    worksheet_rows = resolver.list_children(xlsx_with_object_features, "xlsx:sheet:Data", child_type="row", limit=2)
    range_cells = resolver.list_children(xlsx_with_object_features, "xlsx:sheet:Data!A1:B2", child_type="cell")
    table_rows = resolver.list_children(xlsx_with_object_features, "xlsx:sheet:Data:table:SalesTable", child_type="row")

    assert [child.locator for child in workbook_children] == ["xlsx:sheet:Data", "xlsx:sheet:Archive"]
    assert len(worksheet_rows) == 2
    assert worksheet_rows[0].locator == "xlsx:sheet:Data:row:1"
    assert [child.locator for child in range_cells] == [
        "xlsx:sheet:Data!A1",
        "xlsx:sheet:Data!B1",
        "xlsx:sheet:Data!A2",
        "xlsx:sheet:Data!B2",
    ]
    assert [child.locator for child in table_rows] == ["xlsx:sheet:Data:row:1", "xlsx:sheet:Data:row:2"]


def test_resolve_capabilities_applies_single_sheet_delete_guard(sample_xlsx) -> None:
    resolver = XlsxObjectResolver()

    worksheet_capabilities = resolver.resolve_capabilities(sample_xlsx, "xlsx:sheet:Budget2026")
    formula_capabilities = resolver.resolve_capabilities(sample_xlsx, "xlsx:sheet:Budget2026:formula_cell:C3")

    assert Capability.DELETE in worksheet_capabilities
    assert formula_capabilities == frozenset({Capability.READ, Capability.UPDATE, Capability.STYLE})


def test_xlsx_escape_hatches_write_insert_formula_and_merge(xlsx_with_object_features, tmp_path) -> None:
    write_path = tmp_path / "write.xlsx"
    rows_path = tmp_path / "rows.xlsx"
    cols_path = tmp_path / "cols.xlsx"
    formula_path = tmp_path / "formula.xlsx"
    merge_path = tmp_path / "merge.xlsx"

    written_locator, _, _ = write_range(
        xlsx_with_object_features,
        "xlsx:sheet:Data!A1:B2",
        [["Name", "Value"], ["Beta", "25"]],
        output_path=write_path,
    )
    inserted_row_locator, _, _ = insert_rows(
        xlsx_with_object_features,
        "xlsx:sheet:Data",
        2,
        1,
        output_path=rows_path,
    )
    inserted_col_locator, _, _ = insert_columns(
        xlsx_with_object_features,
        "xlsx:sheet:Data",
        1,
        1,
        output_path=cols_path,
    )
    formula_locator, _, _ = set_formula(
        xlsx_with_object_features,
        "xlsx:sheet:Data!F2",
        "=SUM(B2,5)",
        output_path=formula_path,
    )
    merged_locator, _, _ = merge_cells(
        xlsx_with_object_features,
        "xlsx:sheet:Data!A4:B4",
        output_path=merge_path,
    )

    assert written_locator == "xlsx:sheet:Data!A1:B2"

    workbook = load_workbook(write_path)
    assert workbook["Data"]["A2"].value == "Beta"
    assert workbook["Data"]["B2"].value == 25

    inserted_rows = load_workbook(rows_path)
    assert inserted_row_locator == "xlsx:sheet:Data:row:2"
    assert inserted_rows["Data"]["A3"].value == "Alpha"

    inserted_cols = load_workbook(cols_path)
    assert inserted_col_locator == "xlsx:sheet:Data:col:1"
    assert inserted_cols["Data"]["B1"].value == "Name"

    with_formula = load_workbook(formula_path)
    assert formula_locator == "xlsx:sheet:Data:formula_cell:F2"
    assert with_formula["Data"]["F2"].value == "=SUM(B2,5)"

    merged = load_workbook(merge_path)
    assert merged_locator == "xlsx:sheet:Data:merged_range:A4:B4"
    assert "A4:B4" in {str(cell_range) for cell_range in merged["Data"].merged_cells.ranges}


def test_xlsx_merge_cells_rejects_overlap(xlsx_with_object_features, tmp_path) -> None:
    with pytest.raises(InvalidArgumentsError, match="overlaps existing merged range"):
        merge_cells(
            xlsx_with_object_features,
            "xlsx:sheet:Data!D1:E1",
            output_path=tmp_path / "overlap.xlsx",
        )
