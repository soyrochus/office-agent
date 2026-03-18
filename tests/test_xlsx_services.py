from __future__ import annotations

import pytest
from openpyxl import load_workbook

from offagent.adapters import xlsx_adapter
from offagent.app.services import AppServices, StaleLocatorError
from offagent.config import AppConfig


def test_index_search_locate_read_and_reindex_xlsx(sample_xlsx, tmp_path) -> None:
    services = AppServices(
        AppConfig(index_path=tmp_path / "state" / "index.sqlite3", document_roots=(tmp_path,))
    )

    summary = services.index_path(sample_xlsx)
    hits = services.search_corpus("Supplier shall", file_type="xlsx")
    item = services.locate_cell(sample_xlsx, "Notes 2026", "A1")
    text = services.read_item(sample_xlsx, item.item_id)

    assert summary.files_indexed == 1
    assert hits[0].item_id == "sheet:Notes 2026!A1"
    assert hits[0].preview == "Supplier shall review variance."
    assert item.item_id == "sheet:Notes 2026!A1"
    assert text == "Supplier shall review variance."

    workbook = load_workbook(sample_xlsx)
    workbook["Notes 2026"]["A1"] = "Supplier shall approve variance."
    workbook.save(sample_xlsx)

    services.reindex_path(sample_xlsx)
    updated_hits = services.search_corpus("approve variance", file_type="xlsx")
    assert updated_hits[0].item_id == "sheet:Notes 2026!A1"


def test_write_cell_and_append_xlsx(sample_xlsx, tmp_path) -> None:
    services = AppServices(
        AppConfig(index_path=tmp_path / "state" / "index.sqlite3", document_roots=(tmp_path,))
    )
    services.index_document(sample_xlsx)

    write_result = services.write_cell_value(sample_xlsx, "Budget2026", "B2", "125001")
    append_result = services.append_item_text(
        write_result.output_path,
        "sheet:Notes 2026!B5",
        "Needs review.",
    )

    original_workbook = load_workbook(sample_xlsx)
    written_workbook = load_workbook(write_result.output_path)
    appended_workbook = load_workbook(append_result.output_path)
    assert write_result.text == "125001"
    assert append_result.text == "Needs review."
    assert write_result.output_path != sample_xlsx
    assert ".edited." in write_result.output_path.name
    assert original_workbook["Budget2026"]["B2"].value == 125000
    assert written_workbook["Budget2026"]["B2"].value == 125001
    assert appended_workbook["Notes 2026"]["B5"].value == "Needs review."

    hits = services.search_corpus("Needs review", file_type="xlsx")
    assert hits[0].item_id == "sheet:Notes 2026!B5"
    assert hits[0].document_path == append_result.output_path


def test_append_rejects_numeric_and_formula_xlsx_targets(sample_xlsx, tmp_path) -> None:
    services = AppServices(
        AppConfig(index_path=tmp_path / "state" / "index.sqlite3", document_roots=(tmp_path,))
    )
    services.index_document(sample_xlsx)

    with pytest.raises(xlsx_adapter.TargetNotAppendableError, match="use write-cell"):
        services.append_item_text(sample_xlsx, "sheet:Budget2026!B2", " extra")

    with pytest.raises(xlsx_adapter.TargetNotAppendableError, match="use write-cell"):
        services.append_item_text(sample_xlsx, "sheet:Budget2026!C3", " extra")


def test_write_cell_reresolves_after_external_change_xlsx(sample_xlsx, tmp_path) -> None:
    services = AppServices(
        AppConfig(index_path=tmp_path / "state" / "index.sqlite3", document_roots=(tmp_path,))
    )
    services.index_document(sample_xlsx)

    workbook = load_workbook(sample_xlsx)
    workbook["Notes 2026"]["A1"] = "Supplier shall approve variance."
    workbook.save(sample_xlsx)

    result = services.write_cell_value(sample_xlsx, "Budget2026", "B2", "125001")
    written_workbook = load_workbook(result.output_path)

    assert written_workbook["Budget2026"]["B2"].value == 125001
    assert written_workbook["Notes 2026"]["A1"].value == "Supplier shall approve variance."


def test_write_cell_fails_with_stale_locator_when_target_disappears_xlsx(sample_xlsx, tmp_path) -> None:
    services = AppServices(
        AppConfig(index_path=tmp_path / "state" / "index.sqlite3", document_roots=(tmp_path,))
    )
    services.index_document(sample_xlsx)

    workbook = load_workbook(sample_xlsx)
    del workbook["Budget2026"]
    workbook.save(sample_xlsx)

    with pytest.raises(StaleLocatorError, match="stale locator"):
        services.write_cell_value(sample_xlsx, "Budget2026", "B2", "125001")
