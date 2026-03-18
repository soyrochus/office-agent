from __future__ import annotations

import subprocess
import sys
from pathlib import Path

from openpyxl import load_workbook


def _run_cli(config_path, *args: str) -> subprocess.CompletedProcess[str]:
    return subprocess.run(
        [
            sys.executable,
            "-m",
            "offagent",
            *args,
            "--config",
            str(config_path),
        ],
        capture_output=True,
        text=True,
        check=False,
    )


def _parse_output_path(stdout: str) -> Path:
    return Path(stdout.strip().split("\t")[-1])


def test_xlsx_cli_round_trip(sample_xlsx, config_path) -> None:
    index_result = _run_cli(config_path, "index", str(sample_xlsx))
    search_result = _run_cli(config_path, "search", "Supplier shall", "--type", "xlsx")
    locate_result = _run_cli(
        config_path,
        "locate",
        "--doc",
        str(sample_xlsx),
        "--sheet",
        "Notes 2026",
        "--cell",
        "A1",
    )
    read_result = _run_cli(
        config_path,
        "read",
        "--doc",
        str(sample_xlsx),
        "--item",
        "sheet:Budget2026!C3",
    )
    write_result = _run_cli(
        config_path,
        "write-cell",
        "--doc",
        str(sample_xlsx),
        "--sheet",
        "Budget2026",
        "--cell",
        "B2",
        "--value",
        "125001",
    )
    write_output_path = _parse_output_path(write_result.stdout)
    append_result = _run_cli(
        config_path,
        "append",
        "--doc",
        str(write_output_path),
        "--item",
        "sheet:Notes 2026!B5",
        "--text",
        "CLI added note.",
    )
    append_output_path = _parse_output_path(append_result.stdout)

    assert index_result.returncode == 0, index_result.stderr
    assert "indexed 1" in index_result.stdout
    assert search_result.returncode == 0, search_result.stderr
    assert "sheet:Notes 2026!A1" in search_result.stdout
    assert locate_result.returncode == 0, locate_result.stderr
    assert "sheet:Notes 2026!A1" in locate_result.stdout
    assert read_result.returncode == 0, read_result.stderr
    assert "=SUM(1,2)" in read_result.stdout
    assert write_result.returncode == 0, write_result.stderr
    assert append_result.returncode == 0, append_result.stderr

    original_workbook = load_workbook(sample_xlsx)
    workbook = load_workbook(append_output_path)
    assert original_workbook["Budget2026"]["B2"].value == 125000
    assert workbook["Budget2026"]["B2"].value == 125001
    assert workbook["Notes 2026"]["B5"].value == "CLI added note."


def test_xlsx_cli_append_rejects_numeric_cells(sample_xlsx, config_path) -> None:
    _run_cli(config_path, "index", str(sample_xlsx))

    append_result = _run_cli(
        config_path,
        "append",
        "--doc",
        str(sample_xlsx),
        "--item",
        "sheet:Budget2026!B2",
        "--text",
        "should fail",
    )

    assert append_result.returncode == 1
    assert "write-cell" in append_result.stderr


def test_xlsx_cli_stale_locator_returns_exit_code_3(sample_xlsx, config_path) -> None:
    _run_cli(config_path, "index", str(sample_xlsx))

    workbook = load_workbook(sample_xlsx)
    del workbook["Budget2026"]
    workbook.save(sample_xlsx)

    write_result = _run_cli(
        config_path,
        "write-cell",
        "--doc",
        str(sample_xlsx),
        "--sheet",
        "Budget2026",
        "--cell",
        "B2",
        "--value",
        "125001",
    )

    assert write_result.returncode == 3
    assert "stale locator" in write_result.stderr
