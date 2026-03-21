from __future__ import annotations

import sqlite3

import pytest
from openpyxl import Workbook, load_workbook

from offagent.app.services import AppServices
from offagent.config import AppConfig
from offagent.errors import NoEmbeddingsError


def _vector_config(tmp_path) -> AppConfig:
    return AppConfig(
        index_path=tmp_path / "state" / "index.sqlite3",
        document_roots=(tmp_path,),
        embedding_model="hash://services",
        embedding_dimensions=48,
        vector_search_top_k=10,
    )


def test_index_document_with_embeddings_persists_embedding_rows(sample_docx, tmp_path) -> None:
    services = AppServices(_vector_config(tmp_path))

    services.index_document(sample_docx, with_embeddings=True)

    connection = sqlite3.connect(tmp_path / "state" / "index.sqlite3")
    try:
        row_count = connection.execute("SELECT COUNT(*) FROM item_embeddings").fetchone()[0]
        dimensions = connection.execute(
            "SELECT DISTINCT dimensions FROM item_embeddings"
        ).fetchone()[0]
    finally:
        connection.close()

    assert row_count == 4
    assert dimensions == 48


def test_semantic_and_hybrid_search_return_mode_metadata(sample_docx, tmp_path) -> None:
    services = AppServices(_vector_config(tmp_path))
    services.index_document(sample_docx, with_embeddings=True)

    semantic_hits = services.search_corpus("supplier shall", file_type="docx", mode="semantic")
    hybrid_hits = services.search_corpus("supplier shall", file_type="docx", mode="hybrid")

    assert semantic_hits[0].item_id == "para:3"
    assert semantic_hits[0].match_mode == "semantic"
    assert semantic_hits[0].scores is not None
    assert "semantic" in semantic_hits[0].scores
    assert hybrid_hits[0].item_id == "para:3"
    assert hybrid_hits[0].match_mode == "hybrid"
    assert hybrid_hits[0].scores is not None
    assert {"keyword", "semantic", "final"} <= set(hybrid_hits[0].scores)


def test_semantic_search_requires_indexed_embeddings(sample_docx, tmp_path) -> None:
    services = AppServices(_vector_config(tmp_path))
    services.index_document(sample_docx)

    with pytest.raises(NoEmbeddingsError, match="Reindex with --with-embeddings"):
        services.search_corpus("supplier shall", file_type="docx", mode="semantic")


def test_index_document_with_embeddings_persists_xlsx_row_embedding_rows(sample_xlsx, tmp_path) -> None:
    services = AppServices(_vector_config(tmp_path))

    services.index_document(sample_xlsx, with_embeddings=True)

    connection = sqlite3.connect(tmp_path / "state" / "index.sqlite3")
    try:
        row_count = connection.execute("SELECT COUNT(*) FROM xlsx_row_embeddings").fetchone()[0]
        cell_map_count = connection.execute(
            "SELECT COUNT(*) FROM xlsx_row_embedding_cells"
        ).fetchone()[0]
        item_embedding_count = connection.execute(
            "SELECT COUNT(*) FROM item_embeddings"
        ).fetchone()[0]
    finally:
        connection.close()

    assert row_count == 3
    assert cell_map_count == 3
    assert item_embedding_count == 0


def test_semantic_and_hybrid_search_resolve_xlsx_row_hits_to_cells(sample_xlsx, tmp_path) -> None:
    services = AppServices(_vector_config(tmp_path))
    services.index_document(sample_xlsx, with_embeddings=True)

    semantic_hits = services.search_corpus("supplier shall", file_type="xlsx", mode="semantic")
    hybrid_hits = services.search_corpus("supplier shall", file_type="xlsx", mode="hybrid")

    assert semantic_hits[0].item_id == "sheet:Notes 2026!A1"
    assert semantic_hits[0].match_mode == "semantic"
    assert semantic_hits[0].metadata["matched_sheet"] == "Notes 2026"
    assert semantic_hits[0].metadata["matched_row"] == 1
    assert semantic_hits[0].metadata["contributing_cell_coordinates"] == ["A1"]
    assert semantic_hits[0].metadata["resolved_from_row_embedding"] is True
    assert hybrid_hits[0].item_id == "sheet:Notes 2026!A1"
    assert hybrid_hits[0].match_mode == "hybrid"
    assert hybrid_hits[0].metadata["matched_sheet"] == "Notes 2026"
    assert {"keyword", "semantic", "final"} <= set(hybrid_hits[0].scores or {})


def test_xlsx_semantic_search_requires_indexed_row_embeddings(sample_xlsx, tmp_path) -> None:
    services = AppServices(_vector_config(tmp_path))
    services.index_document(sample_xlsx)

    with pytest.raises(NoEmbeddingsError, match="Reindex with --with-embeddings"):
        services.search_corpus("supplier shall", file_type="xlsx", mode="semantic")


def test_xlsx_row_embeddings_refresh_after_reindex(sample_xlsx, tmp_path) -> None:
    services = AppServices(_vector_config(tmp_path))
    services.index_document(sample_xlsx, with_embeddings=True)

    workbook = load_workbook(sample_xlsx)
    workbook["Notes 2026"]["A1"] = "Budget risk requires approval."
    workbook.save(sample_xlsx)

    services.reindex_path(sample_xlsx, with_embeddings=True)
    semantic_hits = services.search_corpus("approval", file_type="xlsx", mode="semantic")

    connection = sqlite3.connect(tmp_path / "state" / "index.sqlite3")
    try:
        row_count = connection.execute("SELECT COUNT(*) FROM xlsx_row_embeddings").fetchone()[0]
        coordinates = connection.execute(
            "SELECT cell_coordinate FROM xlsx_row_embedding_cells ORDER BY embedding_id, cell_order"
        ).fetchall()
    finally:
        connection.close()

    assert semantic_hits[0].item_id == "sheet:Notes 2026!A1"
    assert row_count == 3
    assert [row[0] for row in coordinates] == ["A1", "A1", "A2"]


def test_xlsx_hybrid_search_uses_resolved_representative_cell_for_row_hits(tmp_path) -> None:
    path = tmp_path / "mixed.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Sheet1"
    worksheet["A1"] = "Status"
    worksheet["B1"] = 125000
    worksheet["C1"] = "Needs review"
    workbook.save(path)

    services = AppServices(_vector_config(tmp_path))
    services.index_document(path, with_embeddings=True)

    hybrid_hits = services.search_corpus("needs review", file_type="xlsx", mode="hybrid")

    assert hybrid_hits[0].item_id == "sheet:Sheet1!C1"
    assert hybrid_hits[0].metadata["contributing_cell_coordinates"] == ["A1", "C1"]
